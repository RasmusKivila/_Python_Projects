from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import os
import pandas as pd

input_path = os.path.join(os.path.dirname(__file__), 'sales_data_v2.csv')
output_path = os.path.join(os.path.dirname(__file__), 'sales_output_v2.xlsx')

class Report():
    def __init__(self, input_path, output_path):
        self.input_path = input_path
        self.output_path = output_path
        self.df = pd.read_csv(input_path)
        
    def calculate_total_sales(self):
        self.df['Total Sales'] = self.df['Count sold'] * self.df['Price per item']
        self.df['Average Price'] = self.df['Total Sales'] / self.df['Count sold']  

    def create_product_sales_sheet(self):
        product_sales = self.df.groupby('Region')['Total Sales'].sum().reset_index()
        product_sales.to_excel(self.output_path, sheet_name='Total sales per region', index=False)
    
    def create_sales_chart(self):
        workbook = Workbook()
        worksheet = workbook.active
        data = Reference(worksheet, min_col=5, min_row=2, max_col=5, max_row=len(self.df) + 1)
        categories = Reference(worksheet, min_col=4, min_row=2, max_row=len(self.df) + 1)
        chart = BarChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = "Sales per month"
        worksheet.add_chart(chart, "E5")
        workbook.save(self.output_path)

    def create_another_sales_sheet(self):
        # Create a new sheet in the existing workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "All Data"

        headers = ['Date', 'Product', 'Region', 'Count sold', 'Price per item']
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)

        for row_num, (_, row) in enumerate(self.df.iterrows(), 2):
            data = [row['Date'], row['Product'], row['Region'], row['Count sold'], row['Price per item']]
            for col_num, value in enumerate(data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Save to the existing Excel file (output_path)
        workbook.save(self.output_path)

    def generate_excel_file(self):
        self.calculate_total_sales()
        self.create_product_sales_sheet()
        self.create_another_sales_sheet()
        self.create_sales_chart()

# Instantiate the class
instance = Report(input_path, output_path)
instance.generate_excel_file()

# TODO

# SPECIFIKATION

# **G**

# - Läs in filen sales_data.csv
# - Använd pandas för att:
#     - Ta bort rader med duplicerade ordernummer
#     - Ta bort rader some har outliers i sitt total_price (över 8000)
#     - Ta bort rader som har tomma product_list och andra eventuella outliers
#     - Konvertera datum som är i annat format till en enhetlig skala
# - Spara ner den till en ny fil kallad cleaned_sales_data.csv

# **VG**

# - Lägg din kod i metoder och lägg dem i en class, instantiera classen och kalla på en metod “clean_data” som i sin tur anropar respektive metod
