import csv
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import os

input_path = os.path.join(os.path.dirname(__file__), 'sales_data.csv')
output_path = os.path.join(os.path.dirname(__file__), 'output.xlsx')

wb = Workbook()

class upg1():
    def create_excel_from_csv(self):
        ws = wb.active
        with open(input_path) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)

    def add_column(self,column_name):
        ws = wb["Sheet"]
        new_column = ws.max_column + 1
        column_count_sold = 4
        column_price_per_item = 5
        
        ws.cell(row = 1, column = new_column, value = column_name)
        
        for row in range(2, ws.max_row+1):
            count_sold = ws.cell(row, column_count_sold).value
            price_per_item = ws.cell(row, column= column_price_per_item).value
            new_column_value = float(count_sold) * float(price_per_item) # Värdena är sparade som string, så gör om dom
            
            ws.cell(row, new_column).value = new_column_value
        
        
    def create_and_populate_new_sheet(self):
        ws1 = wb["Sheet"]
        
        total_sales = {}
        column_products = 2
        column_total_sales = 6
        
        # Summera totala priset för varje unik produkt
        for row in range(2, ws1.max_row+1):
            if ws1.cell(row, column_products).value not in list(total_sales.keys()):
                total_sales[ws1.cell(row, column_products).value] = ws1.cell(row, column_total_sales).value
            else:
                total_sales[ws1.cell(row, column_products).value] += ws1.cell(row, column_total_sales).value
        
        # Spara dessa i ett nytt ark
        wb.create_sheet('Sheet2')
        ws2 = wb["Sheet2"]
        
        ws2.cell(row = 1, column = 1, value = "Products")
        ws2.cell(row = 1, column = 2, value = "Total sales")

        row = 2
        for product in list(total_sales.keys()):
            ws2.cell(row, 1).value = product
            ws2.cell(row, 2).value = total_sales[product]
            row += 1

    def create_bar_chart(self):
        chart = BarChart()
        chart.type = "col"
        chart.title = "Bar Chart"
        chart.x_axis.title = "Products"
        chart.y_axis.title = "Total sales $"
        
        ws2 = wb["Sheet2"]
        data = Reference(ws2, min_col=2, min_row=1, max_col=2, max_row=5)
        products = Reference(ws2, min_col=1, min_row=2, max_row=5)
        # Så du slipper stöta på samma problem nån gång: set_categories() måste vara EFTER add_data() ...
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(products)
        
        ws2.add_chart(chart, "D2")

    def write_csv_data_to_excel(self):
        self.create_excel_from_csv()
        self.add_column('total sales')
        self.create_and_populate_new_sheet()
        self.create_bar_chart()

test = upg1()
test.write_csv_data_to_excel()

wb.save(output_path)


# TODO

# SPECIFIKATION

# **G**

# - Läs in filen sales_data.csv och skriv över datan till en excel-fil kallad output.xlsx
# - Lägg till en kolumn med total försäljning på varje rad
# - Skapa ett nytt ark där du skriver försäljning per produkt
# - Skapa ett stapeldiagram som illustrerar försäljningen per produkt

# **VG**

# - Lägg din kod i metoder och lägg dem i en class, instantiera classen och kalla på en metod “write_csv_data_to_excel” som i sin tur anropar respektive metod